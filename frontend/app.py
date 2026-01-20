import streamlit as st
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.font_manager as fm
from io import BytesIO
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.drawing.fill import SolidColorFillProperties, ColorChoice
from openpyxl.drawing.line import LineProperties
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties, RichTextProperties
from pathlib import Path
import sys
import os
from dotenv import load_dotenv
from templates.template import TEMPLATE_REGISTRY

# Add project root to path for imports (MUST be before template import)
BASE_DIR = Path(__file__).resolve().parent.parent
if str(BASE_DIR) not in sys.path:
    sys.path.insert(0, str(BASE_DIR))


load_dotenv()
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")

# Import chatbot components
from frontend.chatbot.chart_modifier import ChartModifierAgent
from frontend.chatbot.style_updater import StyleUpdater


def main():
    # =====================================================
    # HELPER FUNCTION: Apply VBA-matching Chart Formatting
    # =====================================================
    def apply_vba_formatting(chart, template_style):
        """
        Applies all VBA formatting to openpyxl chart object to match Streamlit preview
        Skips axis formatting for pie charts which don't have axes
        """
        from openpyxl.chart.shapes import GraphicalProperties
        from openpyxl.drawing.text import Font as DrawingFont
        
        def hex_to_rgb(hex_color):
            hex_color = hex_color.lstrip('#')
            return tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))
        
        # 1. REMOVE CHART TITLE ONLY IF NOT PROVIDED
        if not template_style.get("chart_title"):
            chart.title = None
        
        # 2. SET GAP WIDTH = 50% (only for bar/line charts)
        if hasattr(chart, 'gapWidth'):
            chart.gapWidth = 50
        
        # 3. AXIS FORMATTING (skip for pie charts - they don't have axes)
        if not isinstance(chart, PieChart):
            axis_line_rgb = hex_to_rgb(template_style["axis_line_color"])
            axis_text_rgb = hex_to_rgb(template_style["axis_color"])
            
            for axis in [chart.x_axis, chart.y_axis]:
                # Remove gridlines
                axis.majorGridlines = None
                axis.minorGridlines = None
                
                # Format axis line (spine)
                if axis.spPr is None:
                    axis.spPr = GraphicalProperties()
                
                line_props = LineProperties(w=12700)  # Thicker line to match matplotlib
                line_props.solidFill = ColorChoice(
                    srgbClr=f'{axis_line_rgb[0]:02X}{axis_line_rgb[1]:02X}{axis_line_rgb[2]:02X}'
                )
                axis.spPr.ln = line_props
                
                # Tick marks
                axis.majorTickMark = "out"
                axis.minorTickMark = None
                axis.tickLblPos = "low"  # Position labels below/left
                
                # Format axis text (labels)
                if axis.txPr is None:
                    axis.txPr = RichText()
                if axis.txPr.p is None:
                    axis.txPr.p = [Paragraph()]
                
                para = axis.txPr.p[0]
                if para.pPr is None:
                    para.pPr = ParagraphProperties()
                if para.pPr.defRPr is None:
                    para.pPr.defRPr = CharacterProperties()
                
                if para.pPr.defRPr.latin is None:
                    para.pPr.defRPr.latin = DrawingFont(typeface=template_style["font_family"])
                else:
                    para.pPr.defRPr.latin.typeface = template_style["font_family"]
                
                para.pPr.defRPr.sz = template_style["font_size"] * 100
                para.pPr.defRPr.solidFill = ColorChoice(
                    srgbClr=f'{axis_text_rgb[0]:02X}{axis_text_rgb[1]:02X}{axis_text_rgb[2]:02X}'
                )
            
            # Hide top and right spines (like matplotlib does)
            chart.x_axis.spPr = GraphicalProperties() if chart.x_axis.spPr is None else chart.x_axis.spPr
            chart.y_axis.spPr = GraphicalProperties() if chart.y_axis.spPr is None else chart.y_axis.spPr
        
        # 4. LEGEND FORMATTING (works for all chart types)
        if chart.legend:
            legend_text_rgb = hex_to_rgb(template_style["axis_color"])
            chart.legend.position = 'r'  # Position legend on right
            
            if chart.legend.txPr is None:
                chart.legend.txPr = RichText()
            if chart.legend.txPr.p is None:
                chart.legend.txPr.p = [Paragraph()]
            
            para = chart.legend.txPr.p[0]
            if para.pPr is None:
                para.pPr = ParagraphProperties()
            if para.pPr.defRPr is None:
                para.pPr.defRPr = CharacterProperties()
            
            if para.pPr.defRPr.latin is None:
                para.pPr.defRPr.latin = DrawingFont(typeface=template_style["font_family"])
            else:
                para.pPr.defRPr.latin.typeface = template_style["font_family"]
            
            para.pPr.defRPr.sz = template_style["font_size"] * 100
            para.pPr.defRPr.solidFill = ColorChoice(
                srgbClr=f'{legend_text_rgb[0]:02X}{legend_text_rgb[1]:02X}{legend_text_rgb[2]:02X}'
            )
        
        # 5. PLOT AREA - Remove fill to match transparent background
        if hasattr(chart, 'plot_area') and chart.plot_area:
            chart.plot_area.graphicalProperties = GraphicalProperties()
            # Make plot area background transparent
            chart.plot_area.graphicalProperties.noFill = True

    # =====================================================
    # PATH SETUP FOR FONTS
    # =====================================================
    FONTS_DIR = Path(__file__).resolve().parent / "fonts"
    ROBOTO_PATH = FONTS_DIR / "Roboto-Regular.ttf"

    if ROBOTO_PATH.exists():
        fm.fontManager.addfont(str(ROBOTO_PATH))
        plt.rcParams["font.family"] = "Roboto"

    # =====================================================
    # PAGE CONFIG
    # =====================================================
    st.set_page_config(page_title="Excel Branded Charts with AI", layout="wide")
    st.title("📊 Excel Chart Generator with AI Assistant")

    # =====================================================
    # SESSION STATE INITIALIZATION
    # =====================================================
    if 'custom_style' not in st.session_state:
        st.session_state.custom_style = None
    
    if 'chat_history' not in st.session_state:
        st.session_state.chat_history = []
    
    if 'agent' not in st.session_state:
        st.session_state.agent = None

    # =====================================================
    # AI AGENT INITIALIZATION (FROM .env)
    # =====================================================
    if st.session_state.agent is None:
        if not OPENAI_API_KEY:
            st.warning("❌ OPENAI_API_KEY not found in .env")
        else:
            try:
                st.session_state.agent = ChartModifierAgent()
            except Exception as e:
                st.session_state.agent = None
                st.error("❌ Failed to initialize AI Assistant")
                st.exception(e)

    # =====================================================
    # SIDEBAR - TEMPLATE SELECTION
    # =====================================================

    st.sidebar.divider()
    st.sidebar.header("📄 Report Template")
    selected_template = st.sidebar.selectbox(
        "Select Report Template",
        list(TEMPLATE_REGISTRY.keys())
    )
    template_config = TEMPLATE_REGISTRY[selected_template]
    st.sidebar.caption(template_config["description"])

    # Use custom style if available, otherwise use template style
    if st.session_state.custom_style is not None:
        TEMPLATE_STYLE = st.session_state.custom_style
        st.sidebar.info("🎨 Using AI-customized style")
    else:
        TEMPLATE_STYLE = template_config["style"]

    st.sidebar.divider()
    st.sidebar.header("📂 Import & Chart Settings")
    uploaded_file = st.sidebar.file_uploader("Upload Excel", type=["xlsx"])


    # =====================================================
    # MAIN APP
    # =====================================================
    if uploaded_file:
        try:
            df = pd.read_excel(uploaded_file)
            
            # Validate data
            if df.empty:
                st.error("❌ Uploaded file is empty. Please upload a file with data.")
                return
            
            numeric_cols = df.select_dtypes(include=["number"]).columns.tolist()
            all_cols = df.columns.tolist()

            if not numeric_cols:
                st.error("❌ No numeric columns found in the uploaded file.")
                return

            col_x = st.sidebar.selectbox("X-axis (Group By)", all_cols)
            aggregation = st.sidebar.selectbox(
                "Aggregation", ["None", "Sum", "Mean", "Count"]
            )

            default_chart_type = template_config.get("default_chart_type", "Bar Chart")
            chart_type = st.sidebar.selectbox(
                "Chart Type",
                ["Bar Chart", "Line Chart", "Pie Chart"],
                index=["Bar Chart", "Line Chart", "Pie Chart"].index(default_chart_type)
            )

            default_primary = [c for c in template_config.get("default_primary_y", []) if c in numeric_cols]
            default_secondary = [c for c in template_config.get("default_secondary_y", []) if c in numeric_cols]

            if chart_type == "Pie Chart":
                primary_y = [st.sidebar.selectbox(
                    "Y-axis",
                    numeric_cols,
                    index=0 if not default_primary else numeric_cols.index(default_primary[0])
                )]
                secondary_y = []
            else:
                primary_y = st.sidebar.multiselect(
                    "Primary Y-axis",
                    numeric_cols,
                    default=default_primary or numeric_cols[:1]
                )
                secondary_y = st.sidebar.multiselect(
                    "Secondary Y-axis",
                    [c for c in numeric_cols if c not in primary_y],
                    default=default_secondary
                )

            if not primary_y:
                st.error("❌ Please select at least one primary Y-axis column.")
                return

            st.subheader("📝 Edit Data")
            edited_df = st.data_editor(df, width="stretch")

            if aggregation != "None":
                if aggregation == "Sum":
                    plot_df = edited_df.groupby(col_x)[numeric_cols].sum().reset_index()
                elif aggregation == "Mean":
                    plot_df = edited_df.groupby(col_x)[numeric_cols].mean().reset_index()
                elif aggregation == "Count":
                    plot_df = edited_df.groupby(col_x)[numeric_cols].count().reset_index()
            else:
                plot_df = edited_df.copy()

            # Remove NaN values that might occur during aggregation
            plot_df = plot_df.dropna(subset=numeric_cols, how='all')

            # Show current style settings
            with st.expander("🎨 Current Style Settings"):
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.write(f"**Font:** {TEMPLATE_STYLE['font_family']} {TEMPLATE_STYLE['font_size']}pt")
                    st.write(f"**Grid:** {'On' if TEMPLATE_STYLE['grid'] else 'Off'}")
                with col2:
                    st.color_picker("Primary Color", TEMPLATE_STYLE["primary_color"], disabled=True)
                    st.color_picker("Axis Line", TEMPLATE_STYLE["axis_line_color"], disabled=True)
                with col3:
                    st.color_picker("Secondary Color", TEMPLATE_STYLE.get("secondary_color", "#ff7f0e"), disabled=True)
                    st.color_picker("Axis Text", TEMPLATE_STYLE["axis_color"], disabled=True)

            st.subheader("👀 Chart Preview")

            plt.rcParams.update({
                "font.size": TEMPLATE_STYLE["font_size"]
            })

            fig, ax1 = plt.subplots(figsize=(8, 4))
            # Axis titles
            if TEMPLATE_STYLE.get("x_axis_title"):
                ax1.set_xlabel(
                    TEMPLATE_STYLE["x_axis_title"],
                    fontsize=TEMPLATE_STYLE["font_size"],
                    color=TEMPLATE_STYLE["axis_color"]
                )
            else:
                ax1.set_xlabel("")

            if TEMPLATE_STYLE.get("y_axis_title"):
                ax1.set_ylabel(
                    TEMPLATE_STYLE["y_axis_title"],
                    fontsize=TEMPLATE_STYLE["font_size"],
                    color=TEMPLATE_STYLE["axis_color"]
                )
            else:
                ax1.set_ylabel("")
            
            for spine in ["top", "right"]:
                ax1.spines[spine].set_visible(False)
            for spine in ["bottom", "left"]:
                ax1.spines[spine].set_color(TEMPLATE_STYLE["axis_line_color"])
                ax1.spines[spine].set_linewidth(0.5)
            ax1.tick_params(axis="both", colors=TEMPLATE_STYLE["axis_color"], direction="out")

            if TEMPLATE_STYLE.get("chart_title"):
                ax1.set_title(
                    TEMPLATE_STYLE["chart_title"],
                    fontsize=TEMPLATE_STYLE["font_size"] + 2,
                    color=TEMPLATE_STYLE["axis_color"]
                )

            ax2 = ax1.twinx() if secondary_y else None

            if chart_type in ["Bar Chart", "Line Chart"]:
                for col in primary_y:
                    if chart_type == "Bar Chart":
                        ax1.bar(plot_df[col_x], plot_df[col],
                                label=col,
                                color=TEMPLATE_STYLE["primary_color"],
                                alpha=0.85,
                                width=0.5)
                    else:
                        ax1.plot(plot_df[col_x], plot_df[col],
                                 label=col,
                                 linewidth=1.5,
                                 color=TEMPLATE_STYLE["primary_color"])

                if ax2:
                    ax2.spines["top"].set_visible(False)
                    ax2.spines["right"].set_color(TEMPLATE_STYLE["axis_line_color"])
                    ax2.spines["right"].set_linewidth(0.5)
                    ax2.tick_params(colors=TEMPLATE_STYLE["axis_color"], direction="out")
                    for col in secondary_y:
                        ax2.plot(plot_df[col_x], plot_df[col],
                                 linestyle="--",
                                 linewidth=1.2,
                                 label=col,
                                 color=TEMPLATE_STYLE.get("secondary_color", "#ff7f0e"))

                ax1.legend(loc="upper left", frameon=False, fontsize=TEMPLATE_STYLE["font_size"])
                if ax2:
                    ax2.legend(loc="upper right", frameon=False, fontsize=TEMPLATE_STYLE["font_size"])
            else:
                ax1.pie(plot_df[primary_y[0]],
                        labels=plot_df[col_x],
                        autopct="%1.1f%%",
                        textprops={"color": TEMPLATE_STYLE["axis_color"], 
                                  "fontsize": TEMPLATE_STYLE["font_size"]})
                ax1.axis("equal")

            if not TEMPLATE_STYLE["grid"]:
                ax1.grid(False)
            else:
                ax1.grid(True, alpha=0.3)
            
            st.pyplot(fig)
            plt.close(fig)

            st.subheader("⬇️ Export Branded Excel")

            if st.button("Generate Excel Report"):
                wb = Workbook()
                ws = wb.active
                ws.title = template_config.get("sheet_name", "Data")

                number_formats = template_config.get("number_formats", {})

                # Write headers
                for col_idx, col_name in enumerate(plot_df.columns, start=1):
                    cell = ws.cell(row=1, column=col_idx, value=col_name)
                    cell.font = cell.font.copy(bold=True)

                # Write data
                for row_idx, row in enumerate(plot_df.itertuples(index=False), start=2):
                    for col_idx, (col_name, value) in enumerate(zip(plot_df.columns, row), start=1):
                        cell = ws.cell(row=row_idx, column=col_idx, value=value)
                        if col_name in number_formats:
                            cell.number_format = number_formats[col_name]

                end_row = ws.max_row
                end_col = ws.max_column
                table_ref = f"A1:{ws.cell(row=end_row, column=end_col).coordinate}"
                table = Table(displayName="DataTable", ref=table_ref)
                style = TableStyleInfo(
                    name=template_config.get("table_style", "TableStyleMedium9"),
                    showRowStripes=True,
                    showColumnStripes=False
                )
                table.tableStyleInfo = style
                ws.add_table(table)

                # =====================================================
                # CREATE AND FORMAT CHART (FIXED ORDER)
                # =====================================================
                x_idx = 1
                chart_pos = template_config.get("chart_position", "G2")
                # Use the chart_type selected by user, not the template default
                chart_type_to_use = chart_type

                chart = None
                
                if chart_type_to_use in ["Bar Chart", "Line Chart"]:
                    # CREATE CHART FIRST
                    chart = BarChart() if chart_type_to_use == "Bar Chart" else LineChart()
                    
                    # APPLY FORMATTING
                    apply_vba_formatting(chart, TEMPLATE_STYLE)

                    # SET TITLE
                    if TEMPLATE_STYLE.get("chart_title"):
                        chart.title = TEMPLATE_STYLE["chart_title"]
                    else:
                        chart.title = None

                    # Primary axis titles
                    if TEMPLATE_STYLE.get("x_axis_title"):
                        chart.x_axis.title = TEMPLATE_STYLE["x_axis_title"]

                    if TEMPLATE_STYLE.get("y_axis_title"):
                        chart.y_axis.title = TEMPLATE_STYLE["y_axis_title"]

                    # Add primary data
                    for idx, col_name in enumerate(primary_y):
                        y_idx = plot_df.columns.get_loc(col_name) + 1
                        data = Reference(ws, min_col=y_idx, min_row=1, max_row=end_row)
                        chart.add_data(data, titles_from_data=True)
                        
                        if chart.series and idx < len(chart.series):
                            primary_color = TEMPLATE_STYLE["primary_color"].lstrip("#")
                            chart.series[idx].graphicalProperties.solidFill = primary_color

                    # Set categories
                    cats = Reference(ws, min_col=x_idx, min_row=2, max_row=end_row)
                    chart.set_categories(cats)

                    # Add secondary data if exists
                    if secondary_y:
                        sec_chart = LineChart()
                        sec_chart.y_axis.axId = 200
                        
                        if TEMPLATE_STYLE.get("secondary_y_axis_title"):
                            sec_chart.y_axis.title = TEMPLATE_STYLE["secondary_y_axis_title"]
                        
                        for idx, col_name in enumerate(secondary_y):
                            y_idx = plot_df.columns.get_loc(col_name) + 1
                            data = Reference(ws, min_col=y_idx, min_row=1, max_row=end_row)
                            sec_chart.add_data(data, titles_from_data=True)
                            
                            if sec_chart.series and idx < len(sec_chart.series):
                                secondary_color = TEMPLATE_STYLE.get("secondary_color", "#ff7f0e").lstrip("#")
                                sec_chart.series[idx].graphicalProperties.solidFill = secondary_color

                        apply_vba_formatting(sec_chart, TEMPLATE_STYLE)
                        chart += sec_chart
                    
                else:
                    # PIE CHART
                    chart = PieChart()
                    y_idx = plot_df.columns.get_loc(primary_y[0]) + 1
                    data = Reference(ws, min_col=y_idx, min_row=1, max_row=end_row)
                    labels = Reference(ws, min_col=x_idx, min_row=2, max_row=end_row)
                    chart.add_data(data, titles_from_data=True)
                    chart.set_categories(labels)
                    
                    if TEMPLATE_STYLE.get("chart_title"):
                        chart.title = TEMPLATE_STYLE["chart_title"]
                    else:
                        chart.title = None
                    
                    # Add data labels to pie chart
                    if chart.series:
                        from openpyxl.chart.label import DataLabelList
                        chart.series[0].dLbls = DataLabelList()
                        chart.series[0].dLbls.showVal = True
                        chart.series[0].dLbls.showPercent = True
                    
                    # Format legend for pie chart
                    apply_vba_formatting(chart, TEMPLATE_STYLE)

                # Add chart to worksheet
                if chart:
                    ws.add_chart(chart, chart_pos)

                output = BytesIO()
                wb.save(output)
                output.seek(0)

                st.success("✅ Excel file generated with AI-customized styling!")
                
                st.download_button(
                    "⬇️ Download Excel Report",
                    data=output,
                    file_name="ai_customized_excel_report.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

        except Exception as e:
            st.error(f"❌ Error processing file: {str(e)}")
            st.exception(e)

    else:
        st.info("📤 Upload an Excel file from the sidebar to begin.")

    # =====================================================
    # 🤖 AI CHAT ASSISTANT (ALWAYS AT BOTTOM)
    # =====================================================
    st.divider()
    st.subheader("🤖 AI Chart Assistant")
    st.caption("Modify the chart using plain English. Changes apply instantly.")

    if st.session_state.agent is None:
        st.warning("AI Assistant is not available. Check OPENAI_API_KEY.")
    else:
        # ---------- CHAT FORM ----------
        with st.form("chat_form", clear_on_submit=True):
            col1, col2 = st.columns([6, 1])

            with col1:
                user_message = st.text_input(
                    "Type a command (e.g. 'chart title Hello', 'remove grid', 'make font Arial 10')"
                )

            with col2:
                submitted = st.form_submit_button("Send")

            if submitted and user_message.strip():
                current_style = TEMPLATE_STYLE

                modifications = st.session_state.agent.process_command(
                    user_message,
                    current_style
                )

                if modifications:
                    new_style = StyleUpdater.apply_modifications(
                        current_style,
                        modifications
                    )
                    st.session_state.custom_style = new_style

                    st.session_state.chat_history.append(
                        {"role": "user", "content": user_message}
                    )
                    st.session_state.chat_history.append(
                        {
                            "role": "assistant",
                            "content": f"✅ Applied: {', '.join(modifications.keys())}"
                        }
                    )

                    st.rerun()
                else:
                    st.error("❌ I couldn't understand that command.")

        # ---------- CHAT HISTORY ----------
        if st.session_state.chat_history:
            with st.expander("💬 Conversation", expanded=False):
                for msg in st.session_state.chat_history:
                    if msg["role"] == "user":
                        st.markdown(f"**You:** {msg['content']}")
                    else:
                        st.markdown(f"**AI:** {msg['content']}")

        # ---------- CONTROLS (OUTSIDE FORM) ----------
        c1, c2 = st.columns(2)

        with c1:
            if st.button("🔄 Reset Style"):
                st.session_state.custom_style = None
                st.session_state.chat_history = []
                if st.session_state.agent:
                    st.session_state.agent.reset_conversation()
                st.rerun()

        with c2:
            if st.button("🗑️ Clear Chat"):
                st.session_state.chat_history = []
                if st.session_state.agent:
                    st.session_state.agent.reset_conversation()
                st.rerun()

if __name__ == "__main__":
    main()
# src/frontend/excel_exporter.py
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import BarChart, LineChart, PieChart, Reference

# =====================================================
# EXPORT DATA & CHART TO EXCEL
# =====================================================
def export_excel(df: pd.DataFrame, chart_type: str, primary_y: list, secondary_y: list,
                 chart_title: str = "", table_style="TableStyleMedium9", chart_position="G2") -> BytesIO:

    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    # ------------------
    # Write header
    # ------------------
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = cell.font.copy(bold=True)

    # ------------------
    # Write data
    # ------------------
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # ------------------
    # Add Table
    # ------------------
    end_row, end_col = ws.max_row, ws.max_column
    table = Table(displayName="DataTable", ref=f"A1:{ws.cell(end_row, end_col).coordinate}")
    table.tableStyleInfo = TableStyleInfo(name=table_style, showRowStripes=True)
    ws.add_table(table)

    # ------------------
    # Add Chart
    # ------------------
    if chart_type == "Bar Chart":
        chart = BarChart()
    elif chart_type == "Line Chart":
        chart = LineChart()
    else:
        chart = PieChart()

    chart.title = chart_title or chart_type
    for col in primary_y:
        idx = df.columns.get_loc(col) + 1
        data = Reference(ws, min_col=idx, min_row=1, max_row=end_row)
        chart.add_data(data, titles_from_data=True)

    cats = Reference(ws, min_col=1, min_row=2, max_row=end_row)
    chart.set_categories(cats)

    ws.add_chart(chart, chart_position)

    # ------------------
    # Save to BytesIO
    # ------------------
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
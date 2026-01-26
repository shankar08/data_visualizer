"""Excel export functionality"""
from io import BytesIO
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.drawing.fill import SolidColorFillProperties, ColorChoice
from .chart_formatter import apply_vba_formatting


def hex_to_rgb(hex_color):
    """Convert hex color to RGB string for openpyxl"""
    hex_color = hex_color.lstrip('#')
    return hex_color.upper()


def create_excel_report(plot_df, col_x, primary_y, secondary_y, chart_type, 
                       template_config, template_style):
    """
    Create and export Excel workbook with styled chart
    
    Returns: BytesIO object with Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = template_config.get("sheet_name", "Data")

    number_formats = template_config.get("number_formats", {})
    numeric_cols = plot_df.select_dtypes(include=["number"]).columns.tolist()

    # Write headers
    for col_idx, col_name in enumerate(plot_df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = cell.font.copy(bold=True)

    # Write data
    for row_idx, row in enumerate(plot_df.itertuples(index=False), start=2):
        for col_idx, (col_name, value) in enumerate(zip(plot_df.columns, row), start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if col_name in number_formats:
                cell.number_format = number_formats[col_name]

    end_row = ws.max_row
    end_col = ws.max_column
    table_ref = f"A1:{ws.cell(row=end_row, column=end_col).coordinate}"
    table = Table(displayName="DataTable", ref=table_ref)
    style = TableStyleInfo(
        name=template_config.get("table_style", "TableStyleMedium9"),
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    # Create and format chart - PASS template_style here
    chart = create_chart(ws, plot_df, col_x, primary_y, secondary_y, 
                        chart_type, template_style, end_row)
    
    chart_pos = template_config.get("chart_position", "G2")
    if chart:
        ws.add_chart(chart, chart_pos)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def create_chart(ws, plot_df, col_x, primary_y, secondary_y, chart_type, 
                template_style, end_row):
    """Create styled chart based on chart type"""
    chart = None
    x_idx = 1
    
    if chart_type in ["Bar Chart", "Line Chart"]:
        chart = BarChart() if chart_type == "Bar Chart" else LineChart()
        
        # Apply formatting BEFORE setting titles
        apply_vba_formatting(chart, template_style)

        # Chart title
        if template_style.get("chart_title"):
            chart.title = template_style["chart_title"]
        else:
            chart.title = None

        # Axis titles
        if template_style.get("x_axis_title"):
            chart.x_axis.title = template_style["x_axis_title"]

        if template_style.get("y_axis_title"):
            chart.y_axis.title = template_style["y_axis_title"]

        # Add primary data
        for idx, col_name in enumerate(primary_y):
            y_idx = plot_df.columns.get_loc(col_name) + 1
            data = Reference(ws, min_col=y_idx, min_row=1, max_row=end_row)
            chart.add_data(data, titles_from_data=True)
            
            # Apply primary color to series
            if chart.series and idx < len(chart.series):
                primary_color = hex_to_rgb(template_style["primary_color"])
                chart.series[idx].graphicalProperties.solidFill = primary_color

        # Set categories
        cats = Reference(ws, min_col=x_idx, min_row=2, max_row=end_row)
        chart.set_categories(cats)

        # Add secondary data if exists
        if secondary_y:
            sec_chart = LineChart()
            sec_chart.y_axis.axId = 200
            
            if template_style.get("secondary_y_axis_title"):
                sec_chart.y_axis.title = template_style["secondary_y_axis_title"]
            
            for idx, col_name in enumerate(secondary_y):
                y_idx = plot_df.columns.get_loc(col_name) + 1
                data = Reference(ws, min_col=y_idx, min_row=1, max_row=end_row)
                sec_chart.add_data(data, titles_from_data=True)
                
                # Apply secondary color
                if sec_chart.series and idx < len(sec_chart.series):
                    secondary_color = hex_to_rgb(template_style.get("secondary_color", "#ff7f0e"))
                    sec_chart.series[idx].graphicalProperties.solidFill = secondary_color

            apply_vba_formatting(sec_chart, template_style)
            chart += sec_chart
    
    else:
        # PIE CHART
        chart = PieChart()
        y_idx = plot_df.columns.get_loc(primary_y[0]) + 1
        data = Reference(ws, min_col=y_idx, min_row=1, max_row=end_row)
        labels = Reference(ws, min_col=x_idx, min_row=2, max_row=end_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(labels)
        
        if template_style.get("chart_title"):
            chart.title = template_style["chart_title"]
        else:
            chart.title = None
        
        # Add data labels to pie chart
        if chart.series:
            chart.series[0].dLbls = DataLabelList()
            chart.series[0].dLbls.showVal = True
            chart.series[0].dLbls.showPercent = True
        
        apply_vba_formatting(chart, template_style)
    
    return chart